# -*- coding: utf-8 -*-

from argparse import ArgumentParser
from contextlib import closing
from locale import LC_ALL, format, setlocale
from sys import modules

from celery import Celery
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Style
from sqlalchemy import func
from sqlalchemy.orm import backref, relationship

from keyword_analyzer import get_contents
from suggested_keywords import get_suggested_keywords
from top_100_explorer import book, category, trend
from utilities import (
    base,
    get_mysql_session,
    get_popularity,
    get_words,
    variables
)

if 'threading' in modules:
    del modules['threading']

setlocale(LC_ALL, 'en_US.UTF-8')

celery = Celery('hot_keywords')
celery.conf.update(
    BROKER=variables['broker'],
    BROKER_POOL_LIMIT=0,
    CELERY_ACCEPT_CONTENT=['json'],
    CELERY_ACKS_LATE=True,
    CELERY_IGNORE_RESULT=True,
    CELERY_RESULT_SERIALIZER='json',
    CELERY_TASK_SERIALIZER='json',
    CELERYD_LOG_FORMAT='[%(asctime)s: %(levelname)s] %(message)s',
    CELERYD_POOL_RESTARTS=True,
    CELERYD_PREFETCH_MULTIPLIER=1,
    CELERYD_TASK_SOFT_TIME_LIMIT=3600,
    CELERYD_TASK_TIME_LIMIT=7200,
)


class suggested_keyword(base):
    __table_args__ = {
        'autoload': True,
    }
    __tablename__ = 'apps_hot_keywords_suggested_keywords'

    category = relationship(
        'category',
        backref=backref(
            'apps_hot_keywords_suggested_keywords',
            cascade='all,delete-orphan',
            lazy='dynamic',
        ),
    )


class keyword(base):
    __table_args__ = {
        'autoload': True,
    }
    __tablename__ = 'apps_hot_keywords_keywords'


def step_1_1_reset():
    with closing(get_mysql_session()()) as session:
        session.query(suggested_keyword).delete(synchronize_session=False)
        session.commit()


def step_1_1_queue():
    with closing(get_mysql_session()()) as session:
        date = session.query(func.max(trend.date)).first()[0]
        for category_ in session.query(
            category
        ).order_by(
            'id asc'
        ):
            titles = []
            for book_ in session.query(
                book
            ).join(
                trend
            ).filter(
                trend.category_id == category_.id,
                trend.date == date
            ):
                titles.append(book_.title)
            step_1_1_process.delay(
                category_.id, [word for word, _ in get_words(titles, 10)]
            )


@celery.task
def step_1_1_process(category_id, words):
    with closing(get_mysql_session()()) as session:
        for string in get_suggested_keywords(words):
            session.add(suggested_keyword(**{
                'category_id': category_id,
                'string': string,
            }))
            session.commit()


def step_1_2_reset():
    with closing(get_mysql_session()()) as session:
        session.query(
            suggested_keyword
        ).update({
            'popularity': None,
            }, synchronize_session=False
        )
        session.commit()


def step_1_2_queue():
    with closing(get_mysql_session()()) as session:
        for sk in session.query(
            suggested_keyword
        ).filter(
            suggested_keyword.popularity == None
        ).order_by(
            'id asc'
        ):
            step_1_2_process.delay(sk.id, sk.string)


@celery.task
def step_1_2_process(id, string):
    with closing(get_mysql_session()()) as session:
        popularity = get_popularity(string)
        if not popularity:
            return
        sk = session.query(suggested_keyword).get(id)
        sk.popularity = popularity[0]
        session.add(sk)
        session.commit()


def step_2_reset():
    with closing(get_mysql_session()()) as session:
        session.query(keyword).delete(synchronize_session=False)
        session.commit()
        for category_ in session.query(category).order_by('id asc'):
            sk = session.query(
                suggested_keyword
            ).filter(
                suggested_keyword.category_id == category_.id
            ).order_by(
                'popularity desc'
            ).first()
            if sk:
                if not session.query(
                    keyword
                ).filter(
                    keyword.string == sk.string
                ).first():
                    session.add(keyword(**{
                        'string': sk.string,
                    }))
                    session.commit()


def step_2_queue():
    with closing(get_mysql_session()()) as session:
        for k in session.query(
            keyword
        ).filter(
            keyword.score == None
        ).order_by(
            'id asc'
        ):
            step_2_process.delay(k.id, k.string)


@celery.task
def step_2_process(id, string):
    with closing(get_mysql_session()()) as session:
        contents = get_contents(string, 'com')
        if not contents:
            return
        kw = session.query(keyword).get(id)
        kw.count = contents['count'][0]
        kw.buyer_behavior = contents['buyer_behavior'][1]
        kw.competition = contents['competition'][1]
        kw.optimization = contents['optimization'][1]
        kw.popularity = contents['popularity'][1]
        kw.spend = contents['spend'][0][0]
        kw.average_price = contents['average_price'][0]
        kw.average_print_length = contents['average_length'][0]
        kw.score = contents['score'][0]
        session.add(kw)
        session.commit()


def xlsx():
    th_center = Style(
        alignment=Alignment(
            horizontal='center',
            indent=0,
            shrink_to_fit=False,
            text_rotation=0,
            vertical='center',
            wrap_text=False,
        ),
        font=Font(
            bold=True,
            italic=False,
            name='Calibri',
            size=12,
            strike=False,
            underline='none',
            vertAlign=None,
        ),
    )
    th_right = Style(
        alignment=Alignment(
            horizontal='right',
            indent=0,
            shrink_to_fit=False,
            text_rotation=0,
            vertical='center',
            wrap_text=False,
        ),
        font=Font(
            bold=True,
            italic=False,
            name='Calibri',
            size=12,
            strike=False,
            underline='none',
            vertAlign=None,
        ),
    )
    td_center = Style(
        alignment=Alignment(
            horizontal='center',
            indent=0,
            shrink_to_fit=False,
            text_rotation=0,
            vertical='center',
            wrap_text=False,
        ),
        font=Font(
            bold=False,
            italic=False,
            name='Calibri',
            size=10,
            strike=False,
            underline='none',
            vertAlign=None,
        ),
    )
    td_right = Style(
        alignment=Alignment(
            horizontal='right',
            indent=0,
            shrink_to_fit=False,
            text_rotation=0,
            vertical='center',
            wrap_text=False,
        ),
        font=Font(
            bold=False,
            italic=False,
            name='Calibri',
            size=10,
            strike=False,
            underline='none',
            vertAlign=None,
        ),
    )

    workbook = Workbook()

    for worksheet in workbook.worksheets:
        workbook.remove_sheet(worksheet)

    worksheet = workbook.create_sheet()
    worksheet.title = 'Hot Keywords'
    worksheet.cell('A1').style = Style(
        alignment=Alignment(
            horizontal='left',
            indent=0,
            shrink_to_fit=False,
            text_rotation=0,
            vertical='center',
            wrap_text=False,
        ),
        font=Font(
            bold=True,
            italic=False,
            name='Calibri',
            size=12,
            strike=False,
            underline='none',
            vertAlign=None,
        ),
    )
    worksheet.cell('A1').value = 'String'
    worksheet.cell('B1').style = th_right
    worksheet.cell('B1').value = 'Count'
    worksheet.cell('C1').style = th_center
    worksheet.cell('C1').value = 'Buyer Behavior'
    worksheet.cell('D1').style = th_center
    worksheet.cell('D1').value = 'Competition'
    worksheet.cell('E1').style = th_center
    worksheet.cell('E1').value = 'Optimization'
    worksheet.cell('F1').style = th_center
    worksheet.cell('F1').value = 'Popularity'
    worksheet.cell('G1').style = th_right
    worksheet.cell('G1').value = 'Spend'
    worksheet.cell('H1').style = th_right
    worksheet.cell('H1').value = 'Average Price'
    worksheet.cell('I1').style = th_right
    worksheet.cell('I1').value = 'Average Print Length'
    worksheet.cell('J1').style = th_right
    worksheet.cell('J1').value = 'Score'
    worksheet.column_dimensions['A'].width = 50
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 15
    worksheet.column_dimensions['H'].width = 15
    worksheet.column_dimensions['I'].width = 15
    worksheet.column_dimensions['J'].width = 15
    with closing(get_mysql_session()()) as session:
        row = 1
        for kw in session.query(
            keyword
        ).order_by(
            'score desc',
        ).limit(10):
            row += 1
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).value = kw.string
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).style = Style(
                alignment=Alignment(
                    horizontal='left',
                    indent=0,
                    shrink_to_fit=False,
                    text_rotation=0,
                    vertical='center',
                    wrap_text=False,
                ),
                font=Font(
                    bold=False,
                    italic=False,
                    name='Calibri',
                    size=10,
                    strike=False,
                    underline='none',
                    vertAlign=None,
                ),
            )
            worksheet.cell('B%(row)s' % {
                'row': row,
            }).value = format('%d', kw.count, grouping=True)
            worksheet.cell('B%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('C%(row)s' % {
                'row': row,
            }).value = kw.buyer_behavior
            worksheet.cell('C%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('D%(row)s' % {
                'row': row,
            }).value = kw.competition
            worksheet.cell('D%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('E%(row)s' % {
                'row': row,
            }).value = kw.optimization
            worksheet.cell('E%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('F%(row)s' % {
                'row': row,
            }).value = kw.popularity
            worksheet.cell('F%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('G%(row)s' % {
                'row': row,
            }).value = format('%.2f', kw.spend, grouping=True)
            worksheet.cell('G%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('H%(row)s' % {
                'row': row,
            }).value = format(
                '%.2f', kw.average_price, grouping=True,
            )
            worksheet.cell('H%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('I%(row)s' % {
                'row': row,
            }).value = format(
                '%.2f', kw.average_print_length, grouping=True,
            )
            worksheet.cell('I%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('J%(row)s' % {
                'row': row,
            }).value = format('%.2f', kw.score, grouping=True)
            worksheet.cell('J%(row)s' % {
                'row': row,
            }).style = td_right

    workbook.save('../tmp/hot-keywords-report.xlsx')


if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('--def', dest='def_', required=True)

    arguments = parser.parse_args()

    locals()[arguments.def_]()